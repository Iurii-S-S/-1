from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required, current_user
from . import db
from .models import Defect, Project, Comment, Attachment, User
from .utils import role_required, allowed_file
from datetime import datetime
import os
import openpyxl
from io import BytesIO
from sqlalchemy import or_, and_, func

api = Blueprint('api', __name__)

@api.route('/defects', methods=['GET'])
@login_required
def get_defects():
    """Получение списка дефектов с пагинацией и фильтрацией"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 20, type=int), 100)
        
        # Базовый запрос с оптимизацией
        query = Defect.query.options(
            db.joinedload(Defect.creator),
            db.joinedload(Defect.assignee),
            db.joinedload(Defect.project)
        )
        
        # Фильтрация по ролям
        if current_user.role == 'engineer':
            query = query.filter(
                or_(
                    Defect.assignee_id == current_user.id,
                    Defect.creator_id == current_user.id
                )
            )
        elif current_user.role == 'observer':
            query = query.filter(Defect.status.in_(['closed', 'review']))
        
        # Применение фильтров
        filters = []
        
        status = request.args.get('status')
        if status:
            filters.append(Defect.status == status)
        
        priority = request.args.get('priority')
        if priority:
            filters.append(Defect.priority == priority)
        
        project_id = request.args.get('project_id')
        if project_id:
            filters.append(Defect.project_id == project_id)
        
        assignee_id = request.args.get('assignee_id')
        if assignee_id:
            filters.append(Defect.assignee_id == assignee_id)
        
        # Поиск
        search = request.args.get('search')
        if search:
            search_filter = or_(
                Defect.title.ilike(f'%{search}%'),
                Defect.description.ilike(f'%{search}%')
            )
            filters.append(search_filter)
        
        # Просроченные дефекты
        overdue = request.args.get('overdue', type=bool)
        if overdue:
            filters.append(and_(
                Defect.due_date < datetime.utcnow(),
                Defect.status.in_(['new', 'in_progress'])
            ))
        
        if filters:
            query = query.filter(and_(*filters))
        
        # Сортировка
        sort_by = request.args.get('sort_by', 'created_at')
        sort_order = request.args.get('sort_order', 'desc')
        
        sort_mapping = {
            'due_date': Defect.due_date,
            'priority': Defect.priority,
            'status': Defect.status,
            'title': Defect.title,
            'created_at': Defect.created_at,
            'updated_at': Defect.updated_at
        }
        
        order_field = sort_mapping.get(sort_by, Defect.created_at)
        
        if sort_order == 'asc':
            query = query.order_by(order_field.asc())
        else:
            query = query.order_by(order_field.desc())
        
        # Пагинация
        defects = query.paginate(
            page=page, 
            per_page=per_page, 
            error_out=False
        )
        
        return jsonify({
            'success': True,
            'defects': [defect.to_dict() for defect in defects.items],
            'pagination': {
                'total': defects.total,
                'pages': defects.pages,
                'current_page': page,
                'per_page': per_page,
                'has_next': defects.has_next,
                'has_prev': defects.has_prev
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@api.route('/defects/<int:defect_id>', methods=['GET'])
@login_required
def get_defect(defect_id):
    """Получение детальной информации о дефекте"""
    try:
        defect = Defect.query.options(
            db.joinedload(Defect.comments).joinedload(Comment.author),
            db.joinedload(Defect.attachments)
        ).get_or_404(defect_id)
        
        # Проверка прав доступа
        if current_user.role == 'engineer' and defect.assignee_id != current_user.id and defect.creator_id != current_user.id:
            return jsonify({'success': False, 'error': 'Доступ запрещен'}), 403
        elif current_user.role == 'observer' and defect.status not in ['closed', 'review']:
            return jsonify({'success': False, 'error': 'Доступ запрещен'}), 403
        
        defect_data = defect.to_dict()
        defect_data['comments'] = [comment.to_dict() for comment in defect.comments]
        defect_data['attachments'] = [attachment.to_dict() for attachment in defect.attachments]
        
        return jsonify({'success': True, 'defect': defect_data})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@api.route('/defects', methods=['POST'])
@login_required
@role_required(['engineer', 'manager'])
def create_defect():
    """Создание нового дефекта"""
    try:
        data = request.get_json()
        
        if not data or not data.get('title') or not data.get('description'):
            return jsonify({'success': False, 'error': 'Название и описание обязательны'}), 400
        
        # Валидация проекта
        project = Project.query.get(data.get('project_id'))
        if not project:
            return jsonify({'success': False, 'error': 'Проект не найден'}), 400
        
        defect = Defect(
            title=data['title'],
            description=data['description'],
            priority=data.get('priority', 'medium'),
            project_id=data['project_id'],
            creator_id=current_user.id,
            assignee_id=data.get('assignee_id'),
            due_date=datetime.fromisoformat(data['due_date'].replace('Z', '+00:00')) if data.get('due_date') else None
        )
        
        db.session.add(defect)
        db.session.commit()
        
        return jsonify({'success': True, 'defect': defect.to_dict()}), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@api.route('/defects/<int:defect_id>', methods=['PUT'])
@login_required
def update_defect(defect_id):
    """Обновление дефекта"""
    try:
        defect = Defect.query.get_or_404(defect_id)
        data = request.get_json()
        
        # Проверка прав доступа
        if current_user.role != 'manager' and defect.assignee_id != current_user.id:
            return jsonify({'success': False, 'error': 'Доступ запрещен'}), 403
        
        # Обновление полей
        updatable_fields = ['title', 'description', 'priority', 'status', 'assignee_id', 'due_date']
        for field in updatable_fields:
            if field in data:
                if field == 'due_date' and data[field]:
                    setattr(defect, field, datetime.fromisoformat(data[field].replace('Z', '+00:00')))
                else:
                    setattr(defect, field, data[field])
        
        defect.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({'success': True, 'defect': defect.to_dict()})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@api.route('/defects/<int:defect_id>', methods=['DELETE'])
@login_required
@role_required(['manager'])
def delete_defect(defect_id):
    """Удаление дефекта"""
    try:
        defect = Defect.query.get_or_404(defect_id)
        
        db.session.delete(defect)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Дефект удален'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@api.route('/defects/<int:defect_id>/comments', methods=['POST'])
@login_required
def add_comment(defect_id):
    """Добавление комментария к дефекту"""
    try:
        defect = Defect.query.get_or_404(defect_id)
        data = request.get_json()
        
        if not data or not data.get('text'):
            return jsonify({'success': False, 'error': 'Текст комментария обязателен'}), 400
        
        # Проверка прав доступа
        if current_user.role == 'observer':
            return jsonify({'success': False, 'error': 'Наблюдатели не могут оставлять комментарии'}), 403
        
        comment = Comment(
            text=data['text'],
            author_id=current_user.id,
            defect_id=defect_id
        )
        
        db.session.add(comment)
        db.session.commit()
        
        return jsonify({'success': True, 'comment': comment.to_dict()}), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@api.route('/projects', methods=['GET'])
@login_required
def get_projects():
    """Получение списка проектов"""
    try:
        projects = Project.query.filter_by(is_active=True).order_by(Project.name).all()
        
        return jsonify({
            'success': True,
            'projects': [project.to_dict() for project in projects]
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@api.route('/projects', methods=['POST'])
@login_required
@role_required(['manager'])
def create_project():
    """Создание нового проекта"""
    try:
        data = request.get_json()
        
        if not data or not data.get('name'):
            return jsonify({'success': False, 'error': 'Название проекта обязательно'}), 400
        
        project = Project(
            name=data['name'],
            description=data.get('description', ''),
            location=data.get('location', ''),
            start_date=datetime.fromisoformat(data['start_date'].replace('Z', '+00:00')) if data.get('start_date') else None,
            end_date=datetime.fromisoformat(data['end_date'].replace('Z', '+00:00')) if data.get('end_date') else None,
            is_active=data.get('is_active', True)
        )
        
        db.session.add(project)
        db.session.commit()
        
        return jsonify({'success': True, 'project': project.to_dict()}), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@api.route('/statistics/dashboard', methods=['GET'])
@login_required
def get_dashboard_statistics():
    """Статистика для дашборда"""
    try:
        # Основная статистика
        stats_query = db.session.query(
            func.count(Defect.id).label('total'),
            func.count(func.case((Defect.status == 'in_progress', 1))).label('in_progress'),
            func.count(func.case((and_(
                Defect.due_date < datetime.utcnow(),
                Defect.status.in_(['new', 'in_progress'])
            ), 1))).label('overdue'),
            func.count(func.case((Defect.status == 'review', 1))).label('in_review'),
            func.count(func.case((Defect.priority == 'high', 1))).label('high_priority'),
            func.count(func.case((Defect.status == 'closed', 1))).label('closed')
        )
        
        # Фильтрация по ролям
        if current_user.role == 'engineer':
            stats_query = stats_query.filter(
                or_(
                    Defect.assignee_id == current_user.id,
                    Defect.creator_id == current_user.id
                )
            )
        elif current_user.role == 'observer':
            stats_query = stats_query.filter(Defect.status.in_(['closed', 'review']))
        
        stats = stats_query.first()
        
        # Статистика по проектам (только для менеджеров)
        projects_stats = []
        if current_user.role == 'manager':
            projects_data = db.session.query(
                Project.id,
                Project.name,
                func.count(Defect.id).label('total_defects'),
                func.count(func.case((Defect.status == 'in_progress', 1))).label('in_progress'),
                func.count(func.case((Defect.status == 'closed', 1))).label('closed')
            ).outerjoin(Defect).group_by(Project.id, Project.name).all()
            
            projects_stats = [{
                'project_id': project_id,
                'project_name': project_name,
                'total_defects': total_defects,
                'in_progress': in_progress,
                'closed': closed
            } for project_id, project_name, total_defects, in_progress, closed in projects_data]
        
        return jsonify({
            'success': True,
            'statistics': {
                'total_defects': stats.total,
                'in_progress': stats.in_progress,
                'overdue': stats.overdue,
                'in_review': stats.in_review,
                'high_priority': stats.high_priority,
                'closed': stats.closed
            },
            'projects_stats': projects_stats
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@api.route('/export/defects', methods=['GET'])
@login_required
@role_required(['manager'])
def export_defects_to_excel():
    """Экспорт дефектов в Excel"""
    try:
        defects = Defect.query.options(
            db.joinedload(Defect.creator),
            db.joinedload(Defect.assignee),
            db.joinedload(Defect.project)
        ).order_by(Defect.created_at.desc()).all()
        
        # Создание Excel файла
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Дефекты"
        
        # Заголовки
        headers = ['ID', 'Название', 'Описание', 'Статус', 'Приоритет', 'Проект', 
                  'Создатель', 'Исполнитель', 'Срок', 'Дата создания', 'Комментарии']
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        
        # Данные
        for row, defect in enumerate(defects, 2):
            ws.cell(row=row, column=1, value=defect.id)
            ws.cell(row=row, column=2, value=defect.title)
            ws.cell(row=row, column=3, value=defect.description)
            ws.cell(row=row, column=4, value=defect.status)
            ws.cell(row=row, column=5, value=defect.priority)
            ws.cell(row=row, column=6, value=defect.project.name if defect.project else '')
            ws.cell(row=row, column=7, value=defect.creator.first_name + ' ' + defect.creator.last_name)
            ws.cell(row=row, column=8, value=defect.assignee.first_name + ' ' + defect.assignee.last_name if defect.assignee else 'Не назначен')
            ws.cell(row=row, column=9, value=defect.due_date.strftime('%Y-%m-%d') if defect.due_date else '')
            ws.cell(row=row, column=10, value=defect.created_at.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row, column=11, value=len(defect.comments))
        
        # Авто-ширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохранение в память
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'defects_export_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@api.route('/users/engineers', methods=['GET'])
@login_required
def get_engineers():
    """Получение списка инженеров"""
    try:
        engineers = User.query.filter_by(role='engineer').order_by(User.first_name, User.last_name).all()
        
        return jsonify({
            'success': True,
            'engineers': [{
                'id': user.id,
                'name': f'{user.first_name} {user.last_name}',
                'email': user.email
            } for user in engineers]
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500